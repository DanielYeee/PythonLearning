import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel file
file_path = 'your_excel_file.xlsx'
sheet1_name = 'Sheet1'
sheet2_name = 'Sheet2'

# Read the sheets into DataFrames
df1 = pd.read_excel(file_path, sheet_name=sheet1_name)
df2 = pd.read_excel(file_path, sheet_name=sheet2_name)

# Find differences
diff = df1.compare(df2)

# Print differences
print("Differences between the sheets:")
print(diff)

# Highlight differences in the original Excel file
wb = load_workbook(file_path)
ws1 = wb[sheet1_name]
ws2 = wb[sheet2_name]

fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row in range(len(diff)):
    for col in range(len(diff.columns.levels[0])):
        if pd.notna(diff.iloc[row, col]):
            # Highlight in both sheets for visibility
            ws1.cell(row=row+2, column=col+1).fill = fill  # +2 to account for header and 0-based index
            ws2.cell(row=row+2, column=col+1).fill = fill

# Save the modified workbook
output_file = 'highlighted_differences.xlsx'
wb.save(output_file)
print(f"Differences highlighted and saved to {output_file}")
